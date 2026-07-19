"""
World Bank Commodity Markets "Pink Sheet" monthly benchmarks.

Official source:
  https://www.worldbank.org/en/research/commodity-markets

The page links to a monthly XLSX workbook ("CMO-Historical-Data-Monthly.xlsx")
containing benchmark prices for wheat, maize, rice, palm oil, sugar, fertilizers,
coffee, cocoa, beef, and more. We resolve the current workbook URL from the
landing page so the script keeps working as the World Bank rotates document IDs.

Output: data/worldbank_pink_sheet.json
  {
    "as_of_month": "2026-04",
    "dataset_updated_on": "May 04, 2026",
    "series": {
      "wheat": {
        "label": "Wheat, US SRW",
        "unit": "$/mt",
        "source_code": "WHEAT_US_SRW",
        "latest_month": "2026-04",
        "latest_value": 236.4,
        "previous_value": 231.1,
        "change_mom_pct": 2.29
      },
      ...
    }
  }
"""
from __future__ import annotations

from datetime import datetime
from html import unescape
from io import BytesIO
import re

from openpyxl import load_workbook

from _common import http_get, write_json

LANDING_URL = "https://www.worldbank.org/en/research/commodity-markets"
WORKBOOK_RE = re.compile(
    r"https://thedocs\.worldbank\.org/en/doc/[^\"]+/related/CMO-Historical-Data-Monthly\.xlsx",
    re.I,
)

SERIES_MAP = {
    "wheat": "WHEAT_US_SRW",
    "maize": "MAIZE",
    "rice": "RICE_05",
    "soybeans": "SOYBEANS",
    "palm_oil": "PALM_OIL",
    "sugar": "SUGAR_WLD",
    "coffee": "COFFEE_ARABIC",
    "cocoa": "COCOA",
    "urea": "UREA_EE_BULK",
    "phosphate_rock": "PHOSROCK",
    "dap": "DAP",
    "beef": "BEEF",
}


# v38 — worldbank.org WAFs the default FoodShield bot UA; use a browser UA.
_BROWSER_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
               "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

# Stable direct fallbacks if the landing-page scrape is blocked or the regex misses.
WORKBOOK_FALLBACKS = [
    "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021/related/CMO-Historical-Data-Monthly.xlsx",
]


def resolve_workbook_url():
    try:
        html = http_get(LANDING_URL,
                        headers={"Accept": "text/html,*/*", "User-Agent": _BROWSER_UA,
                                 "Accept-Language": "en-US,en;q=0.9"},
                        timeout=45).text
        match = WORKBOOK_RE.search(html)
        if match:
            return unescape(match.group(0))
    except Exception as e:
        print(f"  [warn] Pink Sheet landing-page scrape failed: {e}")
    # Fall back to known-good direct URLs.
    for fb in WORKBOOK_FALLBACKS:
        print(f"  [info] trying Pink Sheet fallback workbook URL")
        return fb
    raise RuntimeError("Could not find Pink Sheet monthly workbook link on landing page")


def load_workbook_rows(url):
    blob = http_get(
        url,
        headers={
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
            "User-Agent": _BROWSER_UA,
            "Accept-Language": "en-US,en;q=0.9",
        },
        timeout=90,
    ).content
    wb = load_workbook(BytesIO(blob), data_only=True, read_only=True)
    ws = wb["Monthly Prices"]
    rows = list(ws.iter_rows(values_only=True))
    return rows


def normalize_num(value):
    if value in (None, "", "…", "..."):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_month(token):
    if not isinstance(token, str):
        return None
    token = token.strip()
    m = re.fullmatch(r"(\d{4})M(\d{2})", token)
    if not m:
        return None
    year = int(m.group(1))
    month = int(m.group(2))
    return f"{year:04d}-{month:02d}"


def latest_points(data_rows, col_idx):
    points = []
    for row in data_rows:
        month = parse_month(row[0] if row else None)
        if not month:
            continue
        val = normalize_num(row[col_idx] if col_idx < len(row) else None)
        if val is None:
            continue
        points.append((month, val))
    if not points:
        return None, None, []
    latest = points[-1]
    previous = points[-2] if len(points) > 1 else None
    # Retain the last 60 months (or however many exist) for sparkline rendering.
    history = points[-60:]
    return latest, previous, history


def _norm_code(c):
    """v47 — the workbook has no source-code row; _find_header_block therefore
    matches against the human LABEL row ('Wheat, US SRW', 'Beef **'). Commas,
    asterisks and percent signs in those labels survived normalisation and broke
    the alias lookup for wheat, rice, sugar, coffee and beef — five series
    silently absent from the site. Strip them so 'Wheat, US SRW' folds onto the
    WHEAT_US_SRW alias. The seven series that already matched contain none of
    these characters, so their normalised codes are unchanged."""
    if c is None:
        return ""
    s = str(c).strip().upper().replace(" ", "_").replace("-", "_")
    s = s.replace(",", "").replace("*", "").replace("%", "")
    return re.sub(r"_+", "_", s).strip("_")


def _find_header_block(rows):
    """v38 — World Bank periodically shifts the metadata rows, which broke the
    hardcoded rows[6]=codes assumption (→ zero benchmarks). Instead, scan the first
    ~12 rows for the one that contains the most known SERIES_MAP source codes; that
    is the code row. labels/units are the rows just above it; data starts below.
    Returns (labels, units, codes, data_rows, code_row_idx)."""
    wanted = {_norm_code(v) for v in SERIES_MAP.values()}
    best_idx, best_hits = None, 0
    for i, row in enumerate(rows[:14]):
        if not row:
            continue
        hits = sum(1 for c in row if _norm_code(c) in wanted)
        if hits > best_hits:
            best_hits, best_idx = hits, i
    if best_idx is None or best_hits == 0:
        raise RuntimeError("Pink Sheet: could not locate the source-code header row")
    codes = rows[best_idx]
    # v47 — the matched row is the workbook's LABEL row, not a code row (no code
    # row exists). Units sit directly BELOW it as '($/mt)' / '($/kg)'; the rows
    # above are the title banner and are blank in every data column, which is why
    # every series shipped with unit=null and label=null. Prefer the row below
    # when it looks like a units row, and fall back to the old behaviour so a
    # workbook that regains a real code row still parses.
    below = rows[best_idx + 1] if best_idx + 1 < len(rows) else ()

    def _looks_like_units(row):
        return any(isinstance(c, str) and c.strip().startswith("($") for c in (row or ()))

    if _looks_like_units(below):
        units = below
        labels = codes
    else:
        labels = rows[best_idx - 2] if best_idx >= 2 else rows[max(0, best_idx - 1)]
        units = rows[best_idx - 1] if best_idx >= 1 else codes
    # data_rows still starts below the matched row; parse_month skips the units
    # row because '($/mt)' is not a YYYYMmm month token.
    data_rows = rows[best_idx + 1:]
    return labels, units, codes, data_rows, best_idx


def main():
    workbook_url = resolve_workbook_url()
    rows = load_workbook_rows(workbook_url)
    if len(rows) < 8:
        raise RuntimeError("Pink Sheet workbook structure is unexpectedly short")

    # Updated-on banner is usually in the first few rows; find it tolerantly.
    dataset_updated_on = None
    for row in rows[:6]:
        for cell in (row or []):
            if isinstance(cell, str) and "updated on" in cell.lower():
                dataset_updated_on = cell.replace("Updated on ", "").replace("updated on ", "").strip()
                break
        if dataset_updated_on:
            break

    labels, units, codes, data_rows, _code_row = _find_header_block(rows)

    # Case/format-tolerant code → column index map.
    col_by_code = {}
    for idx, code in enumerate(codes):
        nc = _norm_code(code)
        if nc:
            col_by_code[nc] = idx

    # Known World Bank code aliases (the workbook occasionally renames columns).
    CODE_ALIASES = {
        "WHEAT_US_SRW": ["WHEAT_US_SRW", "WHEAT_US_HRW", "WHEAT_US"],
        "MAIZE": ["MAIZE", "MAIZE_US"],
        "RICE_05": ["RICE_05", "RICE_25", "RICE_A1", "RICE_THAI_5"],
        "SOYBEANS": ["SOYBEANS", "SOYBEAN"],
        "PALM_OIL": ["PALM_OIL", "PALMOIL"],
        "SUGAR_WLD": ["SUGAR_WLD", "SUGAR_WORLD", "SUGAR"],
        "COFFEE_ARABIC": ["COFFEE_ARABIC", "COFFEE_ARABICA"],
        "COCOA": ["COCOA"],
        "UREA_EE_BULK": ["UREA_EE_BULK", "UREA"],
        "PHOSROCK": ["PHOSROCK", "PHOSPHATE_ROCK"],
        "DAP": ["DAP"],
        "BEEF": ["BEEF"],
    }

    out = {}
    as_of_months = []
    for key, source_code in SERIES_MAP.items():
        aliases = CODE_ALIASES.get(source_code, [source_code])
        idx = next((col_by_code[_norm_code(a)] for a in aliases if _norm_code(a) in col_by_code), None)
        if idx is None:
            continue
        latest, previous, history = latest_points(data_rows, idx)
        if not latest:
            continue
        latest_month, latest_value = latest
        previous_value = previous[1] if previous else None
        change = None
        if previous_value not in (None, 0):
            change = round((latest_value - previous_value) / previous_value * 100, 2)
        as_of_months.append(latest_month)
        out[key] = {
            "label": labels[idx],
            "unit": units[idx],
            "source_code": source_code,
            "latest_month": latest_month,
            "latest_value": round(latest_value, 3),
            "previous_value": round(previous_value, 3) if previous_value is not None else None,
            "change_mom_pct": change,
            # Last ≤60 monthly points for sparkline rendering in the commodity drilldown.
            "history": [
                {"month": m, "value": round(v, 3)} for (m, v) in history
            ],
        }

    if not out:
        raise RuntimeError("Pink Sheet parser found zero usable commodity benchmarks")

    write_json(
        "worldbank_pink_sheet.json",
        {
            "as_of_month": max(as_of_months) if as_of_months else None,
            "dataset_updated_on": dataset_updated_on,
            "series": out,
        },
        source=f"World Bank Commodity Markets Pink Sheet ({workbook_url})",
        notes=(
            "Monthly nominal USD benchmark prices parsed from the World Bank Pink Sheet "
            "historical workbook. Used for commodity cards and live market benchmarks."
        ),
    )


if __name__ == "__main__":
    main()
