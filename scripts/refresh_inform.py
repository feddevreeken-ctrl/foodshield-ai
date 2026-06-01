"""
EU JRC INFORM Risk Index — v23 (Jun 2026): read from committed local file.

The JRC dropped the stable versioned download URL (every
INFORM_Risk_YYYY_vXXX.xlsx path now 404s, and the HDX/WB mirrors are stale).
Rather than chase a moving URL, we commit the official 2026 workbook to the
repo at scripts/vendor_data/INFORM_Risk_2026_v072.xlsx and parse it locally.
This is honest and stable: INFORM publishes annually, so a committed snapshot
is current for a year. To update: download the next edition from
https://drmkc.jrc.ec.europa.eu/inform-index/INFORM-Risk/Results-and-data,
drop it in scripts/vendor_data/, and bump LOCAL_FILE below.

Sheet 'INFORM Risk 2026 (a-z)' layout (data starts row 4, 1-indexed):
  col 0  COUNTRY        col 1  ISO3           col 2  INFORM RISK (0-10)
  col 3  RISK CLASS     col 4  Rank           col 6  HAZARD & EXPOSURE
  col 18 VULNERABILITY  col 30 LACK OF COPING CAPACITY

OUTPUT: data/inform_risk.json
"""
from datetime import datetime, timezone
from pathlib import Path

from _common import write_json

LOCAL_FILE = Path(__file__).resolve().parent / "vendor_data" / "INFORM_Risk_2026_v072.xlsx"
SHEET = "INFORM Risk 2026 (a-z)"
EDITION_YEAR = 2026
EDITION_LABEL = "EU JRC INFORM Risk Index 2026 (v0.7.2)"

# Column indices (0-based) in the (a-z) sheet
COL = {
    "country": 0, "iso3": 1, "inform_risk": 2, "risk_class": 3, "rank": 4,
    "hazard_exposure": 6, "vulnerability": 18, "lack_coping_capacity": 30,
}
DATA_START_ROW = 4  # 1-indexed first country row


def main():
    try:
        import openpyxl
    except ImportError:
        write_json("inform_risk.json", {}, source=EDITION_LABEL,
                   notes="openpyxl not installed; cannot parse XLSX")
        return

    if not LOCAL_FILE.exists():
        write_json("inform_risk.json", {}, source=EDITION_LABEL,
                   notes=(f"Committed workbook missing: {LOCAL_FILE.name}. Download the "
                          f"latest INFORM Risk edition and place it in scripts/vendor_data/."))
        return

    wb = openpyxl.load_workbook(LOCAL_FILE, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        write_json("inform_risk.json", {}, source=EDITION_LABEL,
                   notes=f"Sheet '{SHEET}' not found. Sheets: {wb.sheetnames}")
        return
    ws = wb[SHEET]

    out = {}
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if not row or len(row) <= COL["lack_coping_capacity"]:
            continue
        iso3 = str(row[COL["iso3"]]).strip().upper() if row[COL["iso3"]] else ""
        if len(iso3) != 3 or not iso3.isalpha():
            continue
        risk = _num(row[COL["inform_risk"]])
        if risk is None:
            continue
        out[iso3] = {
            "country": _str(row[COL["country"]]),
            "inform_risk": risk,
            "risk_class": _str(row[COL["risk_class"]]),
            "rank": _int(row[COL["rank"]]),
            "hazard_exposure": _num(row[COL["hazard_exposure"]]),
            "vulnerability": _num(row[COL["vulnerability"]]),
            "lack_coping_capacity": _num(row[COL["lack_coping_capacity"]]),
            "year": EDITION_YEAR,
            "source": EDITION_LABEL,
            "source_url": "https://drmkc.jrc.ec.europa.eu/inform-index/INFORM-Risk",
            "quality_flag": "sourced",
        }

    print(f"[INFO] INFORM: parsed {len(out)} countries from {LOCAL_FILE.name}")
    for ref in ("AFG", "SOM", "YEM", "SDN", "UKR", "USA", "NLD"):
        if ref in out:
            p = out[ref]
            print(f"  [ref] {ref}: risk={p['inform_risk']} ({p['risk_class']}), "
                  f"hazard={p['hazard_exposure']}, vuln={p['vulnerability']}, "
                  f"cope={p['lack_coping_capacity']}")

    write_json(
        "inform_risk.json", out,
        source=EDITION_LABEL,
        notes=(
            f"Composite humanitarian risk 0-10 + 3 dimensions (Hazard & Exposure, "
            f"Vulnerability, Lack of Coping Capacity). Parsed from the official 2026 "
            f"workbook committed at scripts/vendor_data/{LOCAL_FILE.name} (JRC dropped "
            f"the stable download URL). Covered {len(out)} countries. Higher = greater risk. "
            f"Update annually when JRC publishes a new edition."
        ),
    )


def _num(v):
    if v in (None, "", "x", "X", "N.D.", "n.d.", "NA"):
        return None
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return None


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _str(v):
    if v is None or str(v).strip() in ("", "x", "X", "N.D."):
        return None
    return str(v).strip()


if __name__ == "__main__":
    main()
